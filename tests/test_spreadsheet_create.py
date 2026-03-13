"""Tests for create_spreadsheet tool."""

from pathlib import Path

from stratagem.tools.spreadsheet_create import create_spreadsheet


class TestCreateSpreadsheet:
    async def test_basic_creation(self, tmp_path):
        output = tmp_path / "test.xlsx"
        result = await create_spreadsheet.handler({
            "output_path": str(output),
            "sheets": [{
                "name": "Data",
                "headers": ["Name", "Value"],
                "rows": [["Apple", 150], ["Google", 180]],
            }],
        })

        assert output.exists()
        assert "isError" not in result
        assert "2 data rows" in result["content"][0]["text"]

    async def test_multi_sheet(self, tmp_path):
        output = tmp_path / "multi.xlsx"
        result = await create_spreadsheet.handler({
            "output_path": str(output),
            "sheets": [
                {
                    "name": "Revenue",
                    "headers": ["Year", "Amount"],
                    "rows": [["2024", 100], ["2025", 120]],
                },
                {
                    "name": "Expenses",
                    "headers": ["Year", "Amount"],
                    "rows": [["2024", 80], ["2025", 90]],
                },
            ],
        })

        assert output.exists()
        assert "2 sheet(s)" in result["content"][0]["text"]

        # Verify with openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert wb.sheetnames == ["Revenue", "Expenses"]
        ws = wb["Revenue"]
        assert ws.cell(1, 1).value == "Year"
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(2, 2).value == 100
        wb.close()

    async def test_empty_sheets_error(self, tmp_path):
        output = tmp_path / "empty.xlsx"
        result = await create_spreadsheet.handler({
            "output_path": str(output),
            "sheets": [],
        })

        assert result.get("isError") is True
        assert "At least one sheet" in result["content"][0]["text"]
