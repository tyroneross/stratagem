"""Spreadsheet creation tool — create Excel files from structured data."""

from typing import Any
from pathlib import Path

from claude_agent_sdk import tool


@tool(
    "create_spreadsheet",
    "Create an Excel spreadsheet from structured data. Supports multiple sheets with headers and rows.",
    {
        "type": "object",
        "properties": {
            "output_path": {
                "type": "string",
                "description": "Where to save the .xlsx file",
            },
            "sheets": {
                "type": "array",
                "description": "Sheet definitions",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Sheet name"},
                        "headers": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Column headers",
                        },
                        "rows": {
                            "type": "array",
                            "items": {
                                "type": "array",
                                "items": {},
                            },
                            "description": "Data rows (each row is an array of values)",
                        },
                    },
                    "required": ["name", "headers", "rows"],
                },
            },
        },
        "required": ["output_path", "sheets"],
    },
)
async def create_spreadsheet(args: dict[str, Any]) -> dict[str, Any]:
    """Create an Excel workbook with one or more sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _error("openpyxl not installed. Run: pip install openpyxl")

    output_path = args["output_path"]
    sheets = args["sheets"]

    if not sheets:
        return _error("At least one sheet is required")

    # Ensure output directory exists
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet — we'll create named ones
    wb.remove(wb.active)

    total_rows = 0
    for sheet_def in sheets:
        name = sheet_def["name"][:31]  # Excel sheet name limit
        headers = sheet_def["headers"]
        rows = sheet_def["rows"]

        ws = wb.create_sheet(title=name)

        # Write headers with bold font
        bold = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                if col_idx <= len(headers):  # Don't exceed header count
                    ws.cell(row=row_idx, column=col_idx, value=value)

        total_rows += len(rows)

        # Auto-fit column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            # Cap width at 50, add padding
            adjusted = min(max_len + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    wb.save(output_path)

    saved = Path(output_path).resolve()
    if not saved.exists():
        return _error(f"Save completed but file not found at: {saved}")
    size = saved.stat().st_size
    if size == 0:
        return _error(f"File created but is empty: {saved}")

    # Register with artifacts system
    try:
        from stratagem.artifacts import register
        register(
            path=output_path,
            format="xlsx",
            title=Path(output_path).stem,
            cwd=Path.cwd(),
        )
    except Exception:
        pass  # Artifact registration is best-effort

    return {
        "content": [{
            "type": "text",
            "text": f"Spreadsheet saved: {saved} ({len(sheets)} sheet(s), {total_rows} data rows, {size:,} bytes)",
        }]
    }


def _error(msg: str) -> dict[str, Any]:
    return {"content": [{"type": "text", "text": f"Error: {msg}"}], "isError": True}
