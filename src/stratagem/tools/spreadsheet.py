"""Spreadsheet parsing tool — Excel and CSV files to markdown tables."""

from typing import Any
import csv
import io
from pathlib import Path

from claude_agent_sdk import tool


@tool(
    "read_spreadsheet",
    "Read an Excel (.xlsx) or CSV file and return its contents as markdown tables with metadata.",
    {
        "type": "object",
        "properties": {
            "file_path": {"type": "string", "description": "Absolute path to the spreadsheet file"},
            "sheets": {"type": "string", "description": "Comma-separated sheet names or 'all'", "default": "all"},
            "max_rows": {"type": "integer", "description": "Maximum rows per sheet to return", "default": 500},
        },
        "required": ["file_path"],
    },
)
async def read_spreadsheet(args: dict[str, Any]) -> dict[str, Any]:
    file_path = args["file_path"]
    sheets_spec = args.get("sheets", "all")
    max_rows = args.get("max_rows", 500)

    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")

    ext = path.suffix.lower()

    if ext == ".csv":
        return _parse_csv(path, max_rows)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(path, sheets_spec, max_rows)
    else:
        return _error(f"Unsupported format: {ext}. Supported: .xlsx, .xls, .csv")


def _parse_csv(path: Path, max_rows: int) -> dict[str, Any]:
    try:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            dialect = csv.Sniffer().sniff(f.read(8192))
            f.seek(0)
            reader = csv.reader(f, dialect)
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows + 1:  # +1 for header
                    break
                rows.append(row)
    except csv.Error:
        # Fallback to default dialect
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows + 1:
                    break
                rows.append(row)
    except Exception as e:
        return _error(f"CSV parse error: {e}")

    if not rows:
        return _error("CSV file is empty")

    total_rows = len(rows) - 1  # Exclude header
    sections = [
        f"# CSV: {path.name}",
        f"**Rows**: {total_rows} (showing up to {max_rows})",
        f"**Columns**: {len(rows[0])}",
        "",
        _rows_to_markdown(rows),
    ]

    return {"content": [{"type": "text", "text": "\n".join(sections)}]}


def _parse_excel(path: Path, sheets_spec: str, max_rows: int) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _error("openpyxl not installed. Run: pip install openpyxl")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return _error(f"Failed to open workbook: {e}")

    sheet_names = wb.sheetnames
    if sheets_spec == "all":
        target_sheets = sheet_names
    else:
        target_sheets = [s.strip() for s in sheets_spec.split(",")]
        missing = [s for s in target_sheets if s not in sheet_names]
        if missing:
            return _error(f"Sheets not found: {missing}. Available: {sheet_names}")

    sections = [
        f"# Excel: {path.name}",
        f"**Sheets**: {', '.join(sheet_names)}",
        "",
    ]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows + 1:
                break
            rows.append([str(cell if cell is not None else "") for cell in row])

        if not rows:
            sections.append(f"## {sheet_name} (empty)")
            continue

        total_rows = ws.max_row or len(rows)
        total_cols = ws.max_column or (len(rows[0]) if rows else 0)

        sections.append(f"## {sheet_name}")
        sections.append(f"**Rows**: {total_rows} (showing up to {max_rows}) | **Columns**: {total_cols}")
        sections.append("")
        sections.append(_rows_to_markdown(rows))
        sections.append("")

    wb.close()
    return {"content": [{"type": "text", "text": "\n".join(sections)}]}


def _rows_to_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""

    col_count = max(len(r) for r in rows)
    header = rows[0] + [""] * (col_count - len(rows[0]))

    lines = ["| " + " | ".join(header) + " |"]
    lines.append("| " + " | ".join(["---"] * col_count) + " |")

    for row in rows[1:]:
        padded = row + [""] * (col_count - len(row))
        # Escape pipes in cell values
        cleaned = [cell.replace("|", "\\|").replace("\n", " ") for cell in padded]
        lines.append("| " + " | ".join(cleaned) + " |")

    return "\n".join(lines)


def _error(msg: str) -> dict[str, Any]:
    return {"content": [{"type": "text", "text": f"Error: {msg}"}], "isError": True}
